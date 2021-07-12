import pytesseract
import time
import openpyxl
import pyscreenshot
import cv2
from datetime import date
from os import listdir
from os.path import isfile, join
import sys,os

if getattr(sys, 'frozen', False):
    CurrentPath = sys._MEIPASS
else:
    CurrentPath = os.path.dirname(__file__)
print(CurrentPath)
def Capture():
    global i,n
    n=input("Multiple images(y/n): ")
    if n=='n':
        time.sleep(6)
        img=pyscreenshot.grab(bbox=(1579,140,1844,990))
        img.save(CurrentPath+"\\image\\Adc.jpg")
        #img.show()
    else:
        co=1
        i=0
        while co==1:
            co=int(input("Press 1 when ready or 0 to finish: "))
            var="img"+str(i)
            exec("global "+var)
            exec(var+"=pyscreenshot.grab(bbox=(1578,140,1845,990))")
            exec(var+".save(r'"+CurrentPath+'\\image\\'+var+".jpg')")
            i+=1
            #exec(var+".show()")

def ReadImage():
    pytesseract.pytesseract.tesseract_cmd='C:\\Program Files (x86)\\Tesseract-OCR\\tesseract.exe' 
    if n=="y":
        finalli=[]
        for j in range(i):
            var="img"+str(j)
            exec(var+"=cv2.imread(r'"+CurrentPath+'\\image\\'+var+".jpg')")
            exec(var+"=cv2.cvtColor("+var+",cv2.COLOR_BGR2RGB)")
            namess=pytesseract.image_to_string(CurrentPath+'\\image\\'+var+".jpg")
            namesl=namess.split('\n')
            renamesl=[namesl[k] for k in range(0,len(namesl)) if namesl[k]!='' and '(Host)' not in namesl[k]]
            finalli.extend(renamesl)
        result = []
        for name in finalli:
            if name not in result:
                result.append(name)
        file=open(CurrentPath+"\\noteatten\\"+subject+str(date.today())+".txt",'w+')
        file.write(subject+'\n')
        for j in range(len(result)):
            file.write(result[j]+'\n')
        #file.close()
    else:
        img=cv2.imread(CurrentPath+'\\image\\Adc.jpg')
        img=cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
        namess=pytesseract.image_to_string(CurrentPath+"\\image\\Adc.jpg")
        namesl=namess.split('\n')
        fnamesl=[namesl[k] for k in range(0,len(namesl))if namesl[k]!='' and '(Host)' not in namesl[k]]
        file=open(CurrentPath+"\\noteatten\\"+subject+str(date.today())+".txt",'w+')
        file.write(subject+'\n')
        for j in range(len(fnamesl)):
            file.write(fnamesl[j]+'\n')
        #file.close()
        
def AddExcel():
    curedate=str(date.today())
    onlyfiles = [f for f in listdir(CurrentPath+"\\noteatten\\") if isfile(join(CurrentPath+"\\noteatten\\", f))]
    workbook = openpyxl.load_workbook(CurrentPath+'\\AttendenceFinal.xlsx')
    worksheet = workbook.create_sheet(title=curedate)
    worksheet = workbook[curedate]
    row1,col=1,1
    for i in onlyfiles:
        if curedate in i:
            file=open(CurrentPath+"\\noteatten\\"+i,'r')
            file.seek(0)
            filelist=file.readlines()
            subject=filelist[0]
            namelist=filelist[1:]
            worksheet.cell(row=row1,column=col,value=subject)
            for i in namelist:
                worksheet.cell(row=row1+1,column=col,value=i)
                row1 += 1
            col+=1
            row1=1
            file.close()
            continue
    workbook.save(CurrentPath+"\\AttendenceFinal.xlsx")
    workbook.close()            

def main():
    print("1.Capture Image")
    print("2.Read Image")
    print("3.Add to Excel")
    print("4.Quit")
    ch=int(input("Enter your choice: "))
    if ch==1:
        Capture()
        main()
    elif ch==2:
        ReadImage()
        main()
    elif ch==3:
        AddExcel()
        main()
    else:
        exit()

subject=input("Enter the subject name: ")
main()
